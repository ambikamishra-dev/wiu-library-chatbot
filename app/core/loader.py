from __future__ import annotations
from pathlib import Path
from app.config import settings
import openpyxl
from dataclasses import dataclass, field
from typing import Optional
import structlog

log = structlog.get_logger()


@dataclass
class FAQEntry:
    faq_id: str
    question: str
    answer: str

    keywords: list[str] = field(default_factory=list)
    alt_phrasings: list[str] = field(default_factory=list)
    urls: list[dict] = field(default_factory=list)

    category: str = "Other"
    show_as_quick_btn: bool = False
    active: bool = True


def _parse_urls(url_cell: Optional[str], label_cell: Optional[str]) -> list[dict]:
    if not url_cell:
        return []
    urls = [u.strip() for u in str(url_cell).split(";") if u.strip()]
    labels = (
        [l.strip() for l in str(label_cell).split(";")]
        if label_cell
        else []
    )
    result = []
    for i, url in enumerate(urls):
        if not url.startswith("http"):
            url = "https://" + url
        label = labels[i].strip() if i < len(labels) else url
        result.append({"url": url, "label": label})

    return result


def _parse_keywords(raw: Optional[str]) -> list[str]:
    if not raw:
        return []
    return [kw.strip().lower() for kw in str(raw).split(",") if kw.strip()]


def _parse_phrasings(*args) -> list[str]:
    return [
        str(p).strip()
        for p in args
        if p and str(p).strip().lower() not in ("none", "")
    ]


def _parse_bool(value: Optional[str]) -> bool:
    return str(value).strip().upper() == "YES" if value else False


_faq_cache: list[FAQEntry] = []


def load_faq(force_reload: bool = False) -> list[FAQEntry]:
    global _faq_cache
    if _faq_cache and not force_reload:
        return _faq_cache

    path = Path(settings.faq_file_path)

    if not path.exists():
        log.error(f"FAQ file not found", path=str(path))
        raise FileNotFoundError(f"FAQ file not found: {path}")

    if path.suffix.lower() != ".xlsx":
        log.error("faq_invalid_extension", suffix=path.suffix)
        raise ValueError(f"Expected .xlsx file, got: {path.suffix}")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        log.error("faq_file_corrupt", path=str(path), error=str(e))
        raise

    if "FAQ_DATA" not in wb.sheetnames:
        raise KeyError(
            "FAQ_DATA sheet not found in workbook. Check the Excel file.")
    ws = wb["FAQ_DATA"]
    entries: list[FAQEntry] = []

    seen_ids: set[str] = set()

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(cell for cell in row):
            continue

        cells = (list(row) + [None] * 15)[:15]
        (
            faq_id, question, answer, url, url_label,
            keywords, p1, p2, p3, p4,
            category, show_btn, active, notes, last_updated
        ) = cells

        if faq_id and str(faq_id).startswith("Suggested"):
            continue
        if not faq_id or not question or not answer:
            log.warning(
                "faq_row_missing_required_fields",
                faq_id=faq_id,
                has_question=bool(question),
                has_answer=bool(answer),
            )
            continue

        faq_id = str(faq_id).strip()
        if faq_id in seen_ids:
            log.warning("faq_duplicate_id_skipped", faq_id=faq_id)
            continue
        seen_ids.add(faq_id)

        if not _parse_bool(active):
            log.debug("faq_skipped_inactive", faq_id=faq_id)
            continue

        entry = FAQEntry(
            faq_id=faq_id,
            question=str(question).strip(),
            answer=str(answer).strip(),
            keywords=_parse_keywords(keywords),
            alt_phrasings=_parse_phrasings(p1, p2, p3, p4),
            urls=_parse_urls(url, url_label),
            category=str(category).strip() if category else "Other",
            show_as_quick_btn=_parse_bool(show_btn),
            active=True,
        )
        entries.append(entry)

    _faq_cache = entries

    log.info("faq_loaded", count=len(entries), path=str(path))
    return entries


def get_quick_buttons() -> list[FAQEntry]:
    return [e for e in load_faq() if e.show_as_quick_btn]
